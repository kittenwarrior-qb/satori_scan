r"""Nạp dữ liệu hàng loạt từ CodeIT (file Excel báo cáo chi tiết lô) vào DB app.

═══════════════════════════════════════════════════════════════════════════
 CÁCH LẤY DỮ LIỆU RA KHỎI CODEIT
═══════════════════════════════════════════════════════════════════════════
 1. Mở phần mềm CodeIT  ->  màn "QUẢN LÝ CHUNG VÀ BÁO CÁO".
 2. Với MỖI số lô sản xuất (xem cột "Số lô sản xuất" ở bảng trên):
      - Gõ số lô đó vào ô tìm kiếm  ->  bấm "Tìm kiếm"
      - Bấm "Xuất báo cáo"  ->  CodeIT tạo file  .\Data\<SoLoSX>.xlsx
        (chi tiết từng chai: mã chai, số lần tái sử dụng, trạng thái...)
 3. Vào thư mục Data trên máy CodeIT, copy TẤT CẢ file .xlsx vừa xuất
    sang máy này, để chung 1 thư mục (vd: D:\codeit_export\).
 4. Chạy:
        python import_codeit.py D:\codeit_export        (cả thư mục)
        python import_codeit.py D:\codeit_export\STR200108P1.xlsx   (1 file)
        python import_codeit.py D:\codeit_export --tsd 5 --ncc "Ngọc Nghĩa"

 LƯU Ý: file Excel của CodeIT KHÔNG có cột "giới hạn tái sử dụng cho phép".
 Dùng --tsd để đặt giới hạn cho các lô NCC MỚI (mặc định 5). Sau khi import,
 có thể chỉnh lại từng lô ở màn Báo cáo (nút "Cập nhật lô NCC").

 Cột nhận diện (khớp báo cáo CodeIT — Figure 11 trong User Manual):
   Số lô sản xuất | Ngày sản xuất | Số lô nhà cung cấp | Mã số chai
   | Trạng thái chai | Số lần tái sử dụng
═══════════════════════════════════════════════════════════════════════════
"""
import argparse
import os
import sys
from datetime import date, datetime

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import load_workbook

from app.database import models


def _match_columns(headers: list) -> dict:
    """Map tên cột (linh hoạt theo từ khóa) -> chỉ số cột."""
    idx = {}
    for i, h in enumerate(headers):
        h = (str(h) if h is not None else "").strip().lower()
        if "mã" in h and "chai" in h:        idx["ma_chai"] = i
        if "tái sử dụng" in h:               idx["tsd"] = i
        if "trạng thái" in h:                idx["trang_thai"] = i
        if "lô sản xuất" in h:               idx["lo_sx"] = i
        if "nhà cung cấp" in h:              idx["lo_ncc"] = i
        if "ngày" in h:                      idx["ngay"] = i
    return idx


def _parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y",
                "%Y-%m-%d", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_int(v, default=0):
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return default


def import_file(db, path: str, default_tsd: int = 5,
                default_ncc: str = "Ngọc Nghĩa") -> dict:
    """Nạp 1 file .xlsx vào db. Idempotent (chai đã có -> bỏ qua). Trả thống kê."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return {"file": os.path.basename(path), "error": "file rỗng"}

    col = _match_columns(list(headers))
    if "ma_chai" not in col or "lo_sx" not in col:
        return {"file": os.path.basename(path),
                "error": "không tìm thấy cột Mã chai / Lô sản xuất"}

    stats = {"file": os.path.basename(path), "chai_moi": 0, "chai_bo_qua": 0,
             "lo_sx_moi": 0, "lo_ncc_moi": 0}
    sup_cache, prod_cache = {}, {}

    def get_supplier(so_lo_ncc):
        if not so_lo_ncc:
            return None
        if so_lo_ncc in sup_cache:
            return sup_cache[so_lo_ncc]
        s = (db.query(models.SupplierBatch)
             .filter(models.SupplierBatch.so_lo_ncc == so_lo_ncc).first())
        if s is None:
            s = models.SupplierBatch(nha_cung_cap=default_ncc,
                                     so_lo_ncc=so_lo_ncc, so_luong_chai=0,
                                     so_lan_tai_su_dung=default_tsd)
            db.add(s); db.commit(); db.refresh(s)
            stats["lo_ncc_moi"] += 1
        sup_cache[so_lo_ncc] = s
        return s

    def get_production(so_lo_sx, ngay, sup):
        if so_lo_sx in prod_cache:
            return prod_cache[so_lo_sx]
        p = (db.query(models.ProductionBatch)
             .filter(models.ProductionBatch.so_lo_san_xuat == so_lo_sx).first())
        if p is None:
            p = models.ProductionBatch(
                supplier_batch_id=sup.id if sup else None,
                so_lo_san_xuat=so_lo_sx, ngay_san_xuat=ngay or date.today(),
                counter=0)
            db.add(p); db.commit(); db.refresh(p)
            stats["lo_sx_moi"] += 1
        prod_cache[so_lo_sx] = p
        return p

    for row in rows:
        ma_chai = row[col["ma_chai"]] if col["ma_chai"] < len(row) else None
        ma_chai = str(ma_chai).strip() if ma_chai is not None else ""
        if not ma_chai:
            continue

        lo_sx = str(row[col["lo_sx"]]).strip() if row[col["lo_sx"]] else ""
        lo_ncc = (str(row[col["lo_ncc"]]).strip()
                  if "lo_ncc" in col and row[col["lo_ncc"]] else "")
        ngay = _parse_date(row[col["ngay"]]) if "ngay" in col else None
        tsd_thuc_te = _to_int(row[col["tsd"]]) if "tsd" in col else 0
        trang_thai = _to_int(row[col["trang_thai"]]) if "trang_thai" in col else 0

        if db.query(models.Bottle).filter(
                models.Bottle.ma_chai == ma_chai).first():
            stats["chai_bo_qua"] += 1
            continue

        sup = get_supplier(lo_ncc)
        prod = get_production(lo_sx, ngay, sup)
        db.add(models.Bottle(
            ma_chai=ma_chai,
            production_batch_id=prod.id,
            supplier_batch_id=sup.id if sup else None,
            so_lan_thuc_te=tsd_thuc_te,
            trang_thai=trang_thai,
            ngay_san_xuat=ngay or prod.ngay_san_xuat,
        ))
        stats["chai_moi"] += 1

    db.commit()
    wb.close()
    return stats


def main():
    ap = argparse.ArgumentParser(description="Nạp dữ liệu CodeIT (Excel) vào DB")
    ap.add_argument("path", help="File .xlsx hoặc thư mục chứa nhiều .xlsx")
    ap.add_argument("--tsd", type=int, default=5,
                    help="Giới hạn tái sử dụng cho lô NCC mới (mặc định 5)")
    ap.add_argument("--ncc", default="Ngọc Nghĩa", help="Tên nhà cung cấp")
    args = ap.parse_args()

    if os.path.isdir(args.path):
        files = [os.path.join(args.path, f) for f in sorted(os.listdir(args.path))
                 if f.lower().endswith(".xlsx") and not f.startswith("~")]
    elif os.path.isfile(args.path):
        files = [args.path]
    else:
        print(f"Không tìm thấy: {args.path}")
        sys.exit(1)

    if not files:
        print("Không có file .xlsx nào.")
        sys.exit(1)

    from app.database.connection import SessionLocal
    db = SessionLocal()
    total = {"chai_moi": 0, "chai_bo_qua": 0, "lo_sx_moi": 0, "lo_ncc_moi": 0}
    try:
        for f in files:
            st = import_file(db, f, default_tsd=args.tsd, default_ncc=args.ncc)
            if st.get("error"):
                print(f"  ✗ {st['file']}: {st['error']}")
                continue
            for k in total:
                total[k] += st.get(k, 0)
            print(f"  ✓ {st['file']}: +{st['chai_moi']} chai "
                  f"(bỏ qua {st['chai_bo_qua']} đã có)")
    finally:
        db.close()

    print("\n── TỔNG KẾT ──")
    print(f"  Lô NCC mới : {total['lo_ncc_moi']}")
    print(f"  Lô SX mới  : {total['lo_sx_moi']}")
    print(f"  Chai mới   : {total['chai_moi']}")
    print(f"  Chai bỏ qua: {total['chai_bo_qua']} (đã có sẵn)")


if __name__ == "__main__":
    main()
