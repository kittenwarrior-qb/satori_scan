"""Xuất báo cáo Excel (tương thích format CodeIT cũ) + báo cáo theo ca/ngày."""
import os
from collections import defaultdict
from datetime import datetime, time, timedelta
from typing import Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.database import models

REPORT_DIR = os.path.join("data", "reports")

CHE_DO_LABEL = {
    "DINH_DANH": "Định danh",
    "PHAN_LOAI": "Phân loại",
    "LOAI_BO":   "Loại bỏ",
}

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, ncols: int):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center")


def shift_summary(db: Session, from_date=None, to_date=None,
                  che_do: Optional[str] = None) -> dict:
    """Tổng hợp theo CA (mỗi phiên sản xuất) trong khoảng ngày.

    Trả về từng ca + thống kê hợp lệ/lỗi/tỉ lệ + phân tích nguyên nhân lỗi.
    """
    q = db.query(models.Session)
    if from_date:
        q = q.filter(models.Session.bat_dau >= datetime.combine(from_date, time.min))
    if to_date:
        q = q.filter(models.Session.bat_dau
                     < datetime.combine(to_date + timedelta(days=1), time.min))
    if che_do:
        q = q.filter(models.Session.che_do == che_do)
    sessions = q.order_by(models.Session.bat_dau.desc()).all()

    # Phân tích nguyên nhân lỗi: gom scan_events theo (session, ket_qua) 1 lần.
    bd: Dict[int, Dict[str, int]] = {}
    ids = [s.id for s in sessions]
    if ids:
        rows = (db.query(models.ScanEvent.session_id, models.ScanEvent.ket_qua,
                         func.count())
                .filter(models.ScanEvent.session_id.in_(ids))
                .group_by(models.ScanEvent.session_id, models.ScanEvent.ket_qua))
        for sid, kq, cnt in rows:
            bd.setdefault(sid, {})[kq] = cnt

    out_rows, tot_ok, tot_err = [], 0, 0
    for s in sessions:
        ok = s.tong_hop_le or 0
        err = s.tong_loi or 0
        tong = ok + err
        tot_ok += ok
        tot_err += err
        out_rows.append({
            "session_id": s.id,
            "ngay": s.bat_dau.strftime("%d/%m/%Y") if s.bat_dau else "",
            "che_do": s.che_do,
            "che_do_label": CHE_DO_LABEL.get(s.che_do, s.che_do),
            "operator": s.operator or "",
            "bat_dau": s.bat_dau.strftime("%H:%M") if s.bat_dau else "",
            "ket_thuc": s.ket_thuc.strftime("%H:%M") if s.ket_thuc else "—",
            "tong_hop_le": ok,
            "tong_loi": err,
            "tong": tong,
            "ty_le_loi": round(err / tong * 100, 1) if tong else 0,
            "breakdown": bd.get(s.id, {}),
        })

    tong_all = tot_ok + tot_err
    return {
        "rows": out_rows,
        "totals": {
            "tong_hop_le": tot_ok, "tong_loi": tot_err, "tong": tong_all,
            "ty_le_loi": round(tot_err / tong_all * 100, 1) if tong_all else 0,
        },
    }


def export_shifts(db: Session, from_date=None, to_date=None,
                  che_do: Optional[str] = None) -> str:
    """Xuất báo cáo theo ca ra data/reports/ShiftReport.xlsx."""
    os.makedirs(REPORT_DIR, exist_ok=True)
    data = shift_summary(db, from_date, to_date, che_do)
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo theo ca"
    headers = ["Ngày", "Ca", "Trưởng ca", "Bắt đầu", "Kết thúc",
               "Hợp lệ", "Lỗi", "Tổng", "Tỉ lệ lỗi (%)",
               "Quá hạn", "Không đọc", "Mã lạ", "Loại thủ công"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in data["rows"]:
        b = r["breakdown"]
        ws.append([
            r["ngay"], r["che_do_label"], r["operator"], r["bat_dau"],
            r["ket_thuc"], r["tong_hop_le"], r["tong_loi"], r["tong"],
            r["ty_le_loi"], b.get("OVER_LIMIT", 0), b.get("NOREAD", 0),
            b.get("UNKNOWN", 0), b.get("REJECTED", 0),
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 13
    path = os.path.join(REPORT_DIR, "ShiftReport.xlsx")
    wb.save(path)
    return path


def export_production_report(db: Session) -> str:
    """Tổng hợp tất cả lô SX -> data/reports/ProductionReport.xlsx."""
    os.makedirs(REPORT_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo sản xuất"

    # Cột khớp đúng báo cáo CodeIT (Figure 10 trong User Manual):
    # Số lô sản xuất | Số lô nhà cung cấp | Số lượng chai | Ngày sản xuất
    headers = ["Số lô sản xuất", "Số lô nhà cung cấp", "Số lượng chai",
               "Ngày sản xuất"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for b in db.query(models.ProductionBatch).order_by(
            models.ProductionBatch.id).all():
        sup = b.supplier_batch
        sl = (db.query(models.Bottle)
              .filter(models.Bottle.production_batch_id == b.id).count())
        ws.append([
            b.so_lo_san_xuat,
            sup.so_lo_ncc if sup else "",
            sl,
            b.ngay_san_xuat.strftime("%d/%m/%Y") if b.ngay_san_xuat else "",
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    path = os.path.join(REPORT_DIR, "ProductionReport.xlsx")
    wb.save(path)
    return path


def export_batch_detail(db: Session, so_lo_san_xuat: str) -> Optional[str]:
    """Chi tiết chai theo 1 lô SX -> data/reports/<SoLoSX>.xlsx."""
    batch = (db.query(models.ProductionBatch)
             .filter(models.ProductionBatch.so_lo_san_xuat == so_lo_san_xuat)
             .first())
    if batch is None:
        return None

    os.makedirs(REPORT_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = so_lo_san_xuat[:31]

    # Cột khớp đúng báo cáo chi tiết CodeIT (Figure 11 trong User Manual):
    # Số lô SX | Ngày SX | Số lô NCC | Mã số chai | Trạng thái chai | Số lần TSD
    headers = ["Số lô sản xuất", "Ngày sản xuất", "Số lô nhà cung cấp",
               "Mã số chai", "Trạng thái chai", "Số lần tái sử dụng"]
    ws.append(headers)
    _style_header(ws, len(headers))

    sup = batch.supplier_batch
    for bot in (db.query(models.Bottle)
                .filter(models.Bottle.production_batch_id == batch.id)
                .order_by(models.Bottle.id).all()):
        ws.append([
            batch.so_lo_san_xuat,
            bot.ngay_san_xuat.strftime("%d/%m/%Y") if bot.ngay_san_xuat else "",
            sup.so_lo_ncc if sup else "",
            bot.ma_chai,
            bot.trang_thai,
            bot.so_lan_thuc_te,
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    safe = "".join(c for c in so_lo_san_xuat if c.isalnum() or c in "-_")
    path = os.path.join(REPORT_DIR, f"{safe}.xlsx")
    wb.save(path)
    return path


def dashboard_summary(db: Session) -> dict:
    """Số liệu tổng quan cho dashboard biểu đồ ở màn Quản lý & Báo cáo.

    "Thời gian xoay vòng trung bình" là ƯỚC TÍNH dựa trên khoảng cách giữa
    các lần quét OK liên tiếp (+ từ lúc định danh tới lần quét OK đầu tiên)
    của cùng 1 mã chai — hệ thống không lưu vị trí/hành trình thực tế nên
    đây chỉ là xấp xỉ, không phải số đo chính xác như hệ thống warehouse.
    """
    total_bottles = db.query(models.Bottle).count()
    total_batches = db.query(models.ProductionBatch).count()
    total_rejected = (db.query(models.Bottle)
                      .filter(models.Bottle.trang_thai < 0).count())
    ty_le_loi = round(total_rejected / total_bottles * 100, 1) if total_bottles else 0

    # ── Phân bố theo số lần tái sử dụng thực tế ──
    reuse_rows = (db.query(models.Bottle.so_lan_thuc_te, func.count())
                  .group_by(models.Bottle.so_lan_thuc_te)
                  .order_by(models.Bottle.so_lan_thuc_te).all())
    by_reuse = [{"so_lan": n if n is not None else 0, "count": c}
                for n, c in reuse_rows]

    # ── Lý do loại ──
    reason_rows = (db.query(models.Bottle.trang_thai, func.count())
                   .filter(models.Bottle.trang_thai < 0)
                   .group_by(models.Bottle.trang_thai).all())
    by_reject_reason = [
        {"reason": models.REJECT_REASONS.get(tt, f"Khác ({tt})"), "count": c}
        for tt, c in reason_rows
    ]

    # ── Theo từng NCC ──
    by_supplier = []
    for sup in db.query(models.SupplierBatch).all():
        tong = (db.query(models.Bottle)
                .filter(models.Bottle.supplier_batch_id == sup.id).count())
        if tong == 0:
            continue
        loi = (db.query(models.Bottle)
               .filter(models.Bottle.supplier_batch_id == sup.id,
                       models.Bottle.trang_thai < 0).count())
        by_supplier.append({
            "nha_cung_cap": sup.nha_cung_cap, "so_lo_ncc": sup.so_lo_ncc,
            "tong": tong, "loi": loi,
            "ty_le_loi": round(loi / tong * 100, 1) if tong else 0,
        })

    # ── Thời gian xoay vòng trung bình (ước tính) ──
    created_at_by_ma = dict(
        db.query(models.Bottle.ma_chai, models.Bottle.created_at).all())
    scan_rows = (db.query(models.ScanEvent.ma_chai, models.ScanEvent.scanned_at)
                 .filter(models.ScanEvent.event_type == "SCAN",
                         models.ScanEvent.ket_qua == "OK")
                 .order_by(models.ScanEvent.ma_chai, models.ScanEvent.scanned_at)
                 .all())
    grouped: Dict[str, list] = defaultdict(list)
    for ma, ts in scan_rows:
        if ts:
            grouped[ma].append(ts)

    deltas = []
    for ma, timestamps in grouped.items():
        prev = created_at_by_ma.get(ma)
        for ts in timestamps:
            if prev:
                deltas.append((ts - prev).total_seconds())
            prev = ts

    avg_cycle_days = round(sum(deltas) / len(deltas) / 86400, 1) if deltas else None

    return {
        "total_bottles": total_bottles,
        "total_batches": total_batches,
        "total_rejected": total_rejected,
        "ty_le_loi": ty_le_loi,
        "by_reuse": by_reuse,
        "by_reject_reason": by_reject_reason,
        "by_supplier": by_supplier,
        "avg_cycle_days": avg_cycle_days,
    }


def _tt_label(v) -> str:
    """Trạng thái chai dạng chữ (giống CodeIT: Hợp lệ / Bị lỗi)."""
    if v is None or v >= 0:
        return "Hợp lệ"
    return "Bị lỗi"


def bottles_by_date(db: Session, from_date=None, to_date=None,
                    limit: int = 1000) -> dict:
    """Danh sách chai theo khoảng NGÀY SẢN XUẤT (xem nhanh, có giới hạn dòng).

    Truy vấn thẳng bảng bottles → chạy được cả trên dữ liệu cũ import từ CodeIT
    (dữ liệu cũ không có 'ca/session' nên báo cáo theo ca không thấy)."""
    base = db.query(models.Bottle)
    if from_date:
        base = base.filter(models.Bottle.ngay_san_xuat >= from_date)
    if to_date:
        base = base.filter(models.Bottle.ngay_san_xuat <= to_date)

    total = base.count()
    err = base.filter(models.Bottle.trang_thai < 0).count()

    rq = (base.options(joinedload(models.Bottle.production_batch),
                       joinedload(models.Bottle.supplier_batch))
          .order_by(models.Bottle.ngay_san_xuat.desc(), models.Bottle.id)
          .limit(limit))
    rows = []
    for bot in rq.all():
        pb, sb = bot.production_batch, bot.supplier_batch
        rows.append({
            "lo_sx": pb.so_lo_san_xuat if pb else "",
            "lo_ncc": sb.so_lo_ncc if sb else "",
            "ma_chai": bot.ma_chai,
            "so_lan_thuc_te": bot.so_lan_thuc_te,
            "trang_thai": bot.trang_thai,
            "ngay_san_xuat": bot.ngay_san_xuat.strftime("%d/%m/%Y")
            if bot.ngay_san_xuat else "",
        })
    return {"total": total, "ok": total - err, "err": err,
            "shown": len(rows), "rows": rows}


def export_bottles_by_date(db: Session, from_date=None, to_date=None) -> str:
    """Xuất Excel TOÀN BỘ chai trong khoảng ngày (giống CodeIT
    ProductionReport_YYYYMMDD_YYYYMMDD.xls)."""
    os.makedirs(REPORT_DIR, exist_ok=True)
    base = db.query(models.Bottle).options(
        joinedload(models.Bottle.production_batch),
        joinedload(models.Bottle.supplier_batch))
    if from_date:
        base = base.filter(models.Bottle.ngay_san_xuat >= from_date)
    if to_date:
        base = base.filter(models.Bottle.ngay_san_xuat <= to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo theo ngày"
    headers = ["Số lô sản xuất", "Số lô nhà cung cấp", "Mã số chai",
               "Số lần tái sử dụng", "Trạng thái", "Ngày sản xuất"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for bot in base.order_by(models.Bottle.ngay_san_xuat,
                             models.Bottle.id).all():
        pb, sb = bot.production_batch, bot.supplier_batch
        ws.append([
            pb.so_lo_san_xuat if pb else "",
            sb.so_lo_ncc if sb else "",
            bot.ma_chai,
            bot.so_lan_thuc_te,
            _tt_label(bot.trang_thai),
            bot.ngay_san_xuat.strftime("%d/%m/%Y") if bot.ngay_san_xuat else "",
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    f = from_date.strftime("%Y%m%d") if from_date else "dau"
    t = to_date.strftime("%Y%m%d") if to_date else "cuoi"
    path = os.path.join(REPORT_DIR, f"BaoCaoTheoNgay_{f}_{t}.xlsx")
    wb.save(path)
    return path
